from openpyxl import Workbook
import pymongo

"""
Excel文件操作
"""

# 从内存导出Excel文件
def export_from_memory(header,data,filename,sheetname='Sheet1'):
    output_workbook = Workbook()
    output_worksheet = output_workbook.create_sheet(sheetname, index=0)
    data.insert(0, header)
    for row in data:
        output_worksheet.append(row)

    output_workbook.save(filename)


# 从MongoDB导出Excel文件
def export_from_mongoDB(client, db, collection, field_mapping, query=None, sort_by=None, sheetname='Sheet1', filename='result.xlsx',auth=None):
    '''
    从MongoDB导出Excel文件
    :param client: MongoDB连接串
    :param db: 库名
    :param collection:表名
    :param field_mapping:字段映射 传入一个有序字典
    :param query: 查询条件
    :param sort_by: 排序条件
    :param sheetname: sheet名
    :param filename: 文件名
    :param auth: 数据库账号密码
    '''
    client = pymongo.MongoClient(client)
    if auth:
        client.admin.authenticate(*auth)

    db = client[db]
    collection = db[collection]

    output_workbook = Workbook()
    output_worksheet = output_workbook.create_sheet(sheetname, index=0)

    header = list(field_mapping.values())
    output_worksheet.append(header)

    query = {} if not query else query
    results = collection.find(query)

    if sort_by:results.sort(sort_by)

    for result in results:
        row = [result[i] for i in field_mapping.keys()]
        output_worksheet.append(row)

    output_workbook.save(filename)


